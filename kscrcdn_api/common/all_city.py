import pprint
import json
import xlrd
import openpyxl

raw_city = {'上海市': {'city': {}, 'code': '310000'},
 '云南省': {'city': {'临沧市': {'code': '530900'},
                  '丽江市': {'code': '530700'},
                  '保山市': {'code': '530500'},
                  '大理白族自治州': {'code': '532900'},
                  '德宏傣族景颇族自治州': {'code': '533100'},
                  '怒江傈僳族自治州': {'code': '533300'},
                  '文山壮族苗族自治州': {'code': '532600'},
                  '昆明市': {'code': '530100'},
                  '昭通市': {'code': '530600'},
                  '普洱市': {'code': '530800'},
                  '曲靖市': {'code': '530300'},
                  '楚雄彝族自治州': {'code': '532300'},
                  '玉溪市': {'code': '530400'},
                  '红河哈尼族彝族自治州': {'code': '532500'},
                  '西双版纳傣族自治州': {'code': '532800'},
                  '迪庆藏族自治州': {'code': '533400'}},
         'code': '530000'},
 '内蒙古自治区': {'city': {'乌兰察布市': {'code': '150900'},
                     '乌海市': {'code': '150300'},
                     '兴安盟': {'code': '152200'},
                     '包头市': {'code': '150200'},
                     '呼伦贝尔市': {'code': '150700'},
                     '呼和浩特市': {'code': '150100'},
                     '巴彦淖尔市': {'code': '150800'},
                     '赤峰市': {'code': '150400'},
                     '通辽市': {'code': '150500'},
                     '鄂尔多斯市': {'code': '150600'},
                     '锡林郭勒盟': {'code': '152500'},
                     '阿拉善盟': {'code': '152900'}},
            'code': '150000'},
 '北京市': {'city': {}, 'code': '110000'},
 '台湾省': {'city': {}, 'code': '710000'},
 '吉林省': {'city': {'吉林市': {'code': '220200'},
                  '四平市': {'code': '220300'},
                  '延边朝鲜族自治州': {'code': '222400'},
                  '松原市': {'code': '220700'},
                  '白城市': {'code': '220800'},
                  '白山市': {'code': '220600'},
                  '辽源市': {'code': '220400'},
                  '通化市': {'code': '220500'},
                  '长春市': {'code': '220100'}},
         'code': '220000'},
 '四川省': {'city': {'乐山市': {'code': '511100'},
                  '内江市': {'code': '511000'},
                  '凉山彝族自治州': {'code': '513400'},
                  '南充市': {'code': '511300'},
                  '宜宾市': {'code': '511500'},
                  '巴中市': {'code': '511900'},
                  '广元市': {'code': '510800'},
                  '广安市': {'code': '511600'},
                  '德阳市': {'code': '510600'},
                  '成都市': {'code': '510100'},
                  '攀枝花市': {'code': '510400'},
                  '泸州市': {'code': '510500'},
                  '甘孜藏族自治州': {'code': '513300'},
                  '眉山市': {'code': '511400'},
                  '绵阳市': {'code': '510700'},
                  '自贡市': {'code': '510300'},
                  '资阳市': {'code': '512000'},
                  '达州市': {'code': '511700'},
                  '遂宁市': {'code': '510900'},
                  '阿坝藏族羌族自治州': {'code': '513200'},
                  '雅安市': {'code': '511800'}},
         'code': '510000'},
 '天津市': {'city': {}, 'code': '120000'},
 '宁夏回族自治区': {'city': {'中卫市': {'code': '640500'},
                      '吴忠市': {'code': '640300'},
                      '固原市': {'code': '640400'},
                      '石嘴山市': {'code': '640200'},
                      '银川市': {'code': '640100'}},
             'code': '640000'},
 '安徽省': {'city': {'亳州市': {'code': '341600'},
                  '六安市': {'code': '341500'},
                  '合肥市': {'code': '340100'},
                  '安庆市': {'code': '340800'},
                  '宣城市': {'code': '341800'},
                  '宿州市': {'code': '341300'},
                  '池州市': {'code': '341700'},
                  '淮北市': {'code': '340600'},
                  '淮南市': {'code': '340400'},
                  '滁州市': {'code': '341100'},
                  '芜湖市': {'code': '340200'},
                  '蚌埠市': {'code': '340300'},
                  '铜陵市': {'code': '340700'},
                  '阜阳市': {'code': '341200'},
                  '马鞍山市': {'code': '340500'},
                  '黄山市': {'code': '341000'}},
         'code': '340000'},
 '山东省': {'city': {'东营市': {'code': '370500'},
                  '临沂市': {'code': '371300'},
                  '威海市': {'code': '371000'},
                  '德州市': {'code': '371400'},
                  '日照市': {'code': '371100'},
                  '枣庄市': {'code': '370400'},
                  '泰安市': {'code': '370900'},
                  '济南市': {'code': '370100'},
                  '济宁市': {'code': '370800'},
                  '淄博市': {'code': '370300'},
                  '滨州市': {'code': '371600'},
                  '潍坊市': {'code': '370700'},
                  '烟台市': {'code': '370600'},
                  '聊城市': {'code': '371500'},
                  '菏泽市': {'code': '371700'},
                  '青岛市': {'code': '370200'}},
         'code': '370000'},
 '山西省': {'city': {'临汾市': {'code': '141000'},
                  '吕梁市': {'code': '141100'},
                  '大同市': {'code': '140200'},
                  '太原市': {'code': '140100'},
                  '忻州市': {'code': '140900'},
                  '晋中市': {'code': '140700'},
                  '晋城市': {'code': '140500'},
                  '朔州市': {'code': '140600'},
                  '运城市': {'code': '140800'},
                  '长治市': {'code': '140400'},
                  '阳泉市': {'code': '140300'}},
         'code': '140000'},
 '广东省': {'city': {'东莞市': {'code': '441900'},
                  '中山市': {'code': '442000'},
                  '云浮市': {'code': '445300'},
                  '佛山市': {'code': '440600'},
                  '广州市': {'code': '440100'},
                  '惠州市': {'code': '441300'},
                  '揭阳市': {'code': '445200'},
                  '梅州市': {'code': '441400'},
                  '汕头市': {'code': '440500'},
                  '汕尾市': {'code': '441500'},
                  '江门市': {'code': '440700'},
                  '河源市': {'code': '441600'},
                  '深圳市': {'code': '440300'},
                  '清远市': {'code': '441800'},
                  '湛江市': {'code': '440800'},
                  '潮州市': {'code': '445100'},
                  '珠海市': {'code': '440400'},
                  '肇庆市': {'code': '441200'},
                  '茂名市': {'code': '440900'},
                  '阳江市': {'code': '441700'},
                  '韶关市': {'code': '440200'}},
         'code': '440000'},
 '广西壮族自治区': {'city': {'北海市': {'code': '450500'},
                      '南宁市': {'code': '450100'},
                      '崇左市': {'code': '451400'},
                      '来宾市': {'code': '451300'},
                      '柳州市': {'code': '450200'},
                      '桂林市': {'code': '450300'},
                      '梧州市': {'code': '450400'},
                      '河池市': {'code': '451200'},
                      '玉林市': {'code': '450900'},
                      '百色市': {'code': '451000'},
                      '贵港市': {'code': '450800'},
                      '贺州市': {'code': '451100'},
                      '钦州市': {'code': '450700'},
                      '防城港市': {'code': '450600'}},
             'code': '450000'},
 '新疆维吾尔自治区': {'city': {'乌鲁木齐市': {'code': '650100'},
                       '伊犁哈萨克自治州': {'code': '654000'},
                       '克孜勒苏柯尔克孜自治州': {'code': '653000'},
                       '克拉玛依市': {'code': '650200'},
                       '博尔塔拉蒙古自治州': {'code': '652700'},
                       '吐鲁番市': {'code': '650400'},
                       '和田地区': {'code': '653200'},
                       '哈密市': {'code': '650500'},
                       '喀什地区': {'code': '653100'},
                       '塔城地区': {'code': '654200'},
                       '巴音郭楞蒙古自治州': {'code': '652800'},
                       '昌吉回族自治州': {'code': '652300'},
                       '阿克苏地区': {'code': '652900'},
                       '阿勒泰地区': {'code': '654300'}},
              'code': '650000'},
 '江苏省': {'city': {'南京市': {'code': '320100'},
                  '南通市': {'code': '320600'},
                  '宿迁市': {'code': '321300'},
                  '常州市': {'code': '320400'},
                  '徐州市': {'code': '320300'},
                  '扬州市': {'code': '321000'},
                  '无锡市': {'code': '320200'},
                  '泰州市': {'code': '321200'},
                  '淮安市': {'code': '320800'},
                  '盐城市': {'code': '320900'},
                  '苏州市': {'code': '320500'},
                  '连云港市': {'code': '320700'},
                  '镇江市': {'code': '321100'}},
         'code': '320000'},
 '江西省': {'city': {'上饶市': {'code': '361100'},
                  '九江市': {'code': '360400'},
                  '南昌市': {'code': '360100'},
                  '吉安市': {'code': '360800'},
                  '宜春市': {'code': '360900'},
                  '抚州市': {'code': '361000'},
                  '新余市': {'code': '360500'},
                  '景德镇市': {'code': '360200'},
                  '萍乡市': {'code': '360300'},
                  '赣州市': {'code': '360700'},
                  '鹰潭市': {'code': '360600'}},
         'code': '360000'},
 '河北省': {'city': {'保定市': {'code': '130600'},
                  '唐山市': {'code': '130200'},
                  '廊坊市': {'code': '131000'},
                  '张家口市': {'code': '130700'},
                  '承德市': {'code': '130800'},
                  '沧州市': {'code': '130900'},
                  '石家庄市': {'code': '130100'},
                  '秦皇岛市': {'code': '130300'},
                  '衡水市': {'code': '131100'},
                  '邢台市': {'code': '130500'},
                  '邯郸市': {'code': '130400'}},
         'code': '130000'},
 '河南省': {'city': {'三门峡市': {'code': '411200'},
                  '信阳市': {'code': '411500'},
                  '南阳市': {'code': '411300'},
                  '周口市': {'code': '411600'},
                  '商丘市': {'code': '411400'},
                  '安阳市': {'code': '410500'},
                  '平顶山市': {'code': '410400'},
                  '开封市': {'code': '410200'},
                  '新乡市': {'code': '410700'},
                  '洛阳市': {'code': '410300'},
                  '漯河市': {'code': '411100'},
                  '濮阳市': {'code': '410900'},
                  '焦作市': {'code': '410800'},
                  '许昌市': {'code': '411000'},
                  '郑州市': {'code': '410100'},
                  '驻马店市': {'code': '411700'},
                  '鹤壁市': {'code': '410600'}},
         'code': '410000'},
 '浙江省': {'city': {'丽水市': {'code': '331100'},
                  '台州市': {'code': '331000'},
                  '嘉兴市': {'code': '330400'},
                  '宁波市': {'code': '330200'},
                  '杭州市': {'code': '330100'},
                  '温州市': {'code': '330300'},
                  '湖州市': {'code': '330500'},
                  '绍兴市': {'code': '330600'},
                  '舟山市': {'code': '330900'},
                  '衢州市': {'code': '330800'},
                  '金华市': {'code': '330700'}},
         'code': '330000'},
 '海南省': {'city': {'三亚市': {'code': '460200'},
                  '三沙市': {'code': '460300'},
                  '儋州市': {'code': '460400'},
                  '海口市': {'code': '460100'}},
         'code': '460000'},
 '湖北省': {'city': {'十堰市': {'code': '420300'},
                  '咸宁市': {'code': '421200'},
                  '孝感市': {'code': '420900'},
                  '宜昌市': {'code': '420500'},
                  '恩施土家族苗族自治州': {'code': '422800'},
                  '武汉市': {'code': '420100'},
                  '荆州市': {'code': '421000'},
                  '荆门市': {'code': '420800'},
                  '襄阳市': {'code': '420600'},
                  '鄂州市': {'code': '420700'},
                  '随州市': {'code': '421300'},
                  '黄冈市': {'code': '421100'},
                  '黄石市': {'code': '420200'}},
         'code': '420000'},
 '湖南省': {'city': {'娄底市': {'code': '431300'},
                  '岳阳市': {'code': '430600'},
                  '常德市': {'code': '430700'},
                  '张家界市': {'code': '430800'},
                  '怀化市': {'code': '431200'},
                  '株洲市': {'code': '430200'},
                  '永州市': {'code': '431100'},
                  '湘潭市': {'code': '430300'},
                  '湘西土家族苗族自治州': {'code': '433100'},
                  '益阳市': {'code': '430900'},
                  '衡阳市': {'code': '430400'},
                  '邵阳市': {'code': '430500'},
                  '郴州市': {'code': '431000'},
                  '长沙市': {'code': '430100'}},
         'code': '430000'},
 '澳门特别行政区': {'city': {}, 'code': '820000'},
 '甘肃省': {'city': {'临夏回族自治州': {'code': '622900'},
                  '兰州市': {'code': '620100'},
                  '嘉峪关市': {'code': '620200'},
                  '天水市': {'code': '620500'},
                  '定西市': {'code': '621100'},
                  '平凉市': {'code': '620800'},
                  '庆阳市': {'code': '621000'},
                  '张掖市': {'code': '620700'},
                  '武威市': {'code': '620600'},
                  '甘南藏族自治州': {'code': '623000'},
                  '白银市': {'code': '620400'},
                  '酒泉市': {'code': '620900'},
                  '金昌市': {'code': '620300'},
                  '陇南市': {'code': '621200'}},
         'code': '620000'},
 '福建省': {'city': {'三明市': {'code': '350400'},
                  '南平市': {'code': '350700'},
                  '厦门市': {'code': '350200'},
                  '宁德市': {'code': '350900'},
                  '泉州市': {'code': '350500'},
                  '漳州市': {'code': '350600'},
                  '福州市': {'code': '350100'},
                  '莆田市': {'code': '350300'},
                  '龙岩市': {'code': '350800'}},
         'code': '350000'},
 '西藏自治区': {'city': {'山南市': {'code': '540500'},
                    '拉萨市': {'code': '540100'},
                    '日喀则市': {'code': '540200'},
                    '昌都市': {'code': '540300'},
                    '林芝市': {'code': '540400'},
                    '那曲市': {'code': '540600'},
                    '阿里地区': {'code': '542500'}},
           'code': '540000'},
 '贵州省': {'city': {'六盘水市': {'code': '520200'},
                  '安顺市': {'code': '520400'},
                  '毕节市': {'code': '520500'},
                  '贵阳市': {'code': '520100'},
                  '遵义市': {'code': '520300'},
                  '铜仁市': {'code': '520600'},
                  '黔东南苗族侗族自治州': {'code': '522600'},
                  '黔南布依族苗族自治州': {'code': '522700'},
                  '黔西南布依族苗族自治州': {'code': '522300'}},
         'code': '520000'},
 '辽宁省': {'city': {'丹东市': {'code': '210600'},
                  '大连市': {'code': '210200'},
                  '抚顺市': {'code': '210400'},
                  '朝阳市': {'code': '211300'},
                  '本溪市': {'code': '210500'},
                  '沈阳市': {'code': '210100'},
                  '盘锦市': {'code': '211100'},
                  '营口市': {'code': '210800'},
                  '葫芦岛市': {'code': '211400'},
                  '辽阳市': {'code': '211000'},
                  '铁岭市': {'code': '211200'},
                  '锦州市': {'code': '210700'},
                  '阜新市': {'code': '210900'},
                  '鞍山市': {'code': '210300'}},
         'code': '210000'},
 '重庆市': {'city': {}, 'code': '500000'},
 '陕西省': {'city': {'咸阳市': {'code': '610400'},
                  '商洛市': {'code': '611000'},
                  '安康市': {'code': '610900'},
                  '宝鸡市': {'code': '610300'},
                  '延安市': {'code': '610600'},
                  '榆林市': {'code': '610800'},
                  '汉中市': {'code': '610700'},
                  '渭南市': {'code': '610500'},
                  '西安市': {'code': '610100'},
                  '铜川市': {'code': '610200'}},
         'code': '610000'},
 '青海省': {'city': {'果洛藏族自治州': {'code': '632600'},
                  '海东市': {'code': '630200'},
                  '海北藏族自治州': {'code': '632200'},
                  '海南藏族自治州': {'code': '632500'},
                  '海西蒙古族藏族自治州': {'code': '632800'},
                  '玉树藏族自治州': {'code': '632700'},
                  '西宁市': {'code': '630100'},
                  '黄南藏族自治州': {'code': '632300'}},
         'code': '630000'},
 '香港特别行政区': {'city': {}, 'code': '810000'},
 '黑龙江省': {'city': {'七台河市': {'code': '230900'},
                   '伊春市': {'code': '230700'},
                   '佳木斯市': {'code': '230800'},
                   '双鸭山市': {'code': '230500'},
                   '哈尔滨市': {'code': '230100'},
                   '大兴安岭地区': {'code': '232700'},
                   '大庆市': {'code': '230600'},
                   '牡丹江市': {'code': '231000'},
                   '绥化市': {'code': '231200'},
                   '鸡西市': {'code': '230300'},
                   '鹤岗市': {'code': '230400'},
                   '黑河市': {'code': '231100'},
                   '齐齐哈尔市': {'code': '230200'}},
          'code': '230000'}}


__city__ = {
    u"广西": {
        "name_pinyin": "GuangXi",
        "name_abbreviation": "gx",
        "code": "45",
        "city": {
            u"防城港": {
                "name_pinyin": "FangChengGang",
                "name_abbreviation": "fcg",
                "code": "4506"
            },
            u"柳州": {
                "name_pinyin": "LiuZhou",
                "name_abbreviation": "liuz",
                "code": "4502"
            },
            u"南宁": {
                "name_pinyin": "NanNing",
                "name_abbreviation": "nn",
                "code": "4501"
            },
            u"梧州": {
                "name_pinyin": "WuZhou",
                "name_abbreviation": "wuzh",
                "code": "4504"
            },
            u"百色": {
                "name_pinyin": "BaiSe",
                "name_abbreviation": "bs",
                "code": "4510"
            },
            u"贺州": {
                "name_pinyin": "HeZhou",
                "name_abbreviation": "hezh",
                "code": "4511"
            },
            u"玉林": {
                "name_pinyin": "YuLin",
                "name_abbreviation": "yul",
                "code": "4509"
            },
            u"北海": {
                "name_pinyin": "BeiHai",
                "name_abbreviation": "bh",
                "code": "4505"
            },
            u"崇左": {
                "name_pinyin": "ChongZuo",
                "name_abbreviation": "chz",
                "code": "4514"
            },
            u"贵港": {
                "name_pinyin": "GuiGang",
                "name_abbreviation": "gg",
                "code": "4508"
            },
            u"桂林": {
                "name_pinyin": "GuiLin",
                "name_abbreviation": "gl",
                "code": "4503"
            },
            u"钦州": {
                "name_pinyin": "QinZhou",
                "name_abbreviation": "qzh",
                "code": "4507"
            },
            u"来宾": {
                "name_pinyin": "LaiBin",
                "name_abbreviation": "lb",
                "code": "4513"
            },
            u"河池": {
                "name_pinyin": "HeChi",
                "name_abbreviation": "hch",
                "code": "4512"
            }
        }
    },
    u"陕西": {
        "name_pinyin": "ShanXi",
        "name_abbreviation": "shx",
        "code": "61",
        "city": {
            u"汉中": {
                "name_pinyin": "HanZhong",
                "name_abbreviation": "hanz",
                "code": "6107"
            },
            u"渭南": {
                "name_pinyin": "WeiNan",
                "name_abbreviation": "wn",
                "code": "6105"
            },
            u"商洛": {
                "name_pinyin": "ShangLuo",
                "name_abbreviation": "shl",
                "code": "6110"
            },
            u"宝鸡": {
                "name_pinyin": "BaoJi",
                "name_abbreviation": "baoj",
                "code": "6103"
            },
            u"西安": {
                "name_pinyin": "XiAn",
                "name_abbreviation": "xa",
                "code": "6101"
            },
            u"延安": {
                "name_pinyin": "YanAn",
                "name_abbreviation": "yanan",
                "code": "6106"
            },
            u"咸阳": {
                "name_pinyin": "XianYang",
                "name_abbreviation": "xiany",
                "code": "6104"
            },
            u"铜川": {
                "name_pinyin": "TongChuan",
                "name_abbreviation": "tc",
                "code": "6102"
            },
            u"榆林": {
                "name_pinyin": "YuLin",
                "name_abbreviation": "yl",
                "code": "6108"
            },
            u"安康": {
                "name_pinyin": "AnKang",
                "name_abbreviation": "ak",
                "code": "6109"
            }
        }
    },
    u"上海": {
        "city": {
            u"上海": {
                "name_pinyin": "ShangHai",
                "name_abbreviation": "sh",
                "code": "310000"
            }
        },
        "name_pinyin": "ShangHai",
        "name_abbreviation": "sh",
        "code": "310000"
    },
    u"河北": {
        "name_pinyin": "HeBei",
        "name_abbreviation": "hb",
        "code": "13",
        "city": {
            u"张家口": {
                "name_pinyin": "ZhangJiaKou",
                "name_abbreviation": "zjk",
                "code": "1307"
            },
            u"唐山": {
                "name_pinyin": "TangShan",
                "name_abbreviation": "ts",
                "code": "1302"
            },
            u"秦皇岛": {
                "name_pinyin": "QinHuangDao",
                "name_abbreviation": "qhd",
                "code": "1303"
            },
            u"承德": {
                "name_pinyin": "ChengDe",
                "name_abbreviation": "chengd",
                "code": "1308"
            },
            u"沧州": {
                "name_pinyin": "CangZhou",
                "name_abbreviation": "cangz",
                "code": "1309"
            },
            u"石家庄": {
                "name_pinyin": "ShiJiaZhuang",
                "name_abbreviation": "sjz",
                "code": "1301"
            },
            u"保定": {
                "name_pinyin": "BaoDing",
                "name_abbreviation": "bd",
                "code": "1306"
            },
            u"邢台": {
                "name_pinyin": "XingTai",
                "name_abbreviation": "xt",
                "code": "1305"
            },
            u"衡水": {
                "name_pinyin": "HengShui",
                "name_abbreviation": "hs",
                "code": "1311"
            },
            u"廊坊": {
                "name_pinyin": "LangFang",
                "name_abbreviation": "langf",
                "code": "1310"
            },
            u"邯郸": {
                "name_pinyin": "HanDan",
                "name_abbreviation": "hd",
                "code": "1304"
            }
        }
    },
    u"湖南": {
        "name_pinyin": "HuNan",
        "name_abbreviation": "hn",
        "code": "43",
        "city": {
            u"郴州": {
                "name_pinyin": "ChenZhou",
                "name_abbreviation": "chenz",
                "code": "4310"
            },
            u"长沙": {
                "name_pinyin": "ChangSha",
                "name_abbreviation": "cs",
                "code": "4301"
            },
            u"湘潭": {
                "name_pinyin": "XiangTan",
                "name_abbreviation": "xiangt",
                "code": "4303"
            },
            u"岳阳": {
                "name_pinyin": "YueYang",
                "name_abbreviation": "yy",
                "code": "4306"
            },
            u"张家界": {
                "name_pinyin": "ZhangJiaJie",
                "name_abbreviation": "zhjj",
                "code": "4308"
            },
            u"株洲": {
                "name_pinyin": "ZhuZhou",
                "name_abbreviation": "zhz",
                "code": "4302"
            },
            u"益阳": {
                "name_pinyin": "YiYang",
                "name_abbreviation": "yiy",
                "code": "4309"
            },
            u"衡阳": {
                "name_pinyin": "HengYang",
                "name_abbreviation": "hy",
                "code": "4304"
            },
            u"永州": {
                "name_pinyin": "YongZhou",
                "name_abbreviation": "yzh",
                "code": "4311"
            },
            u"邵阳": {
                "name_pinyin": "ShaoYang",
                "name_abbreviation": "shy",
                "code": "4305"
            },
            u"怀化": {
                "name_pinyin": "HuaiHua",
                "name_abbreviation": "huaih",
                "code": "4312"
            },
            u"娄底": {
                "name_pinyin": "LouDi",
                "name_abbreviation": "ld",
                "code": "4313"
            },
            u"常德": {
                "name_pinyin": "ChangDe",
                "name_abbreviation": "changd",
                "code": "4307"
            },
            u"湘西": {
                "name_pinyin": "XiangXi",
                "name_abbreviation": "xiangx",
                "code": "4331"
            }
        }
    },
    u"河南": {
        "name_pinyin": "HeNan",
        "name_abbreviation": "hn",
        "code": "41",
        "city": {
            u"洛阳": {
                "name_pinyin": "LuoYang",
                "name_abbreviation": "ly",
                "code": "4103"
            },
            u"开封": {
                "name_pinyin": "KaiFeng",
                "name_abbreviation": "kf",
                "code": "4102"
            },
            u"许昌": {
                "name_pinyin": "XuChang",
                "name_abbreviation": "xc",
                "code": "4110"
            },
            u"郑州": {
                "name_pinyin": "ZhengZhou",
                "name_abbreviation": "zz",
                "code": "4101"
            },
            u"商丘": {
                "name_pinyin": "ShangQiu",
                "name_abbreviation": "shq",
                "code": "4114"
            },
            u"安阳": {
                "name_pinyin": "AnYang",
                "name_abbreviation": "ay",
                "code": "4105"
            },
            u"焦作": {
                "name_pinyin": "JiaoZuo",
                "name_abbreviation": "jiaoz",
                "code": "4108"
            },
            u"鹤壁": {
                "name_pinyin": "HeBi",
                "name_abbreviation": "heb",
                "code": "4106"
            },
            u"济源": {
                "name_pinyin": "JiYuan",
                "name_abbreviation": "jy",
                "code": "4190"
            },
            u"信阳": {
                "name_pinyin": "XinYang",
                "name_abbreviation": "xiny",
                "code": "4115"
            },
            u"新乡": {
                "name_pinyin": "XinXiang",
                "name_abbreviation": "xx",
                "code": "4107"
            },
            u"南阳": {
                "name_pinyin": "NanYang",
                "name_abbreviation": "ny",
                "code": "4113"
            },
            u"三门峡": {
                "name_pinyin": "SanMenXia",
                "name_abbreviation": "smx",
                "code": "4112"
            },
            u"驻马店": {
                "name_pinyin": "ZhuMaDian",
                "name_abbreviation": "zmd",
                "code": "4117"
            },
            u"周口": {
                "name_pinyin": "ZhouKou",
                "name_abbreviation": "zk",
                "code": "4116"
            },
            u"濮阳": {
                "name_pinyin": "PuYang",
                "name_abbreviation": "py",
                "code": "4109"
            },
            u"漯河": {
                "name_pinyin": "luoHe",
                "name_abbreviation": "luoh",
                "code": "4111"
            },
            u"平顶山": {
                "name_pinyin": "PingDingShan",
                "name_abbreviation": "pds",
                "code": "4104"
            }
        }
    },
    u"山东": {
        "name_pinyin": "ShanDong",
        "name_abbreviation": "shd",
        "code": "37",
        "city": {
            u"烟台": {
                "name_pinyin": "YanTai",
                "name_abbreviation": "yt",
                "code": "3706"
            },
            u"滨州": {
                "name_pinyin": "BinZhou",
                "name_abbreviation": "bz",
                "code": "3716"
            },
            u"济宁": {
                "name_pinyin": "JiNing",
                "name_abbreviation": "jin",
                "code": "3708"
            },
            u"日照": {
                "name_pinyin": "RiZhao",
                "name_abbreviation": "rzh",
                "code": "3711"
            },
            u"菏泽": {
                "name_pinyin": "HeZe",
                "name_abbreviation": "hez",
                "code": "3717"
            },
            u"潍坊": {
                "name_pinyin": "WeiFang",
                "name_abbreviation": "wf",
                "code": "3707"
            },
            u"聊城": {
                "name_pinyin": "LiaoCheng",
                "name_abbreviation": "lch",
                "code": "3715"
            },
            u"东营": {
                "name_pinyin": "DongYing",
                "name_abbreviation": "dongy",
                "code": "3705"
            },
            u"淄博": {
                "name_pinyin": "ZiBo",
                "name_abbreviation": "zb",
                "code": "3703"
            },
            u"威海": {
                "name_pinyin": "WeiHai",
                "name_abbreviation": "weih",
                "code": "3710"
            },
            u"青岛": {
                "name_pinyin": "QingDao",
                "name_abbreviation": "qd",
                "code": "3702"
            },
            u"德州": {
                "name_pinyin": "DeZhou",
                "name_abbreviation": "dez",
                "code": "3714"
            },
            u"泰安": {
                "name_pinyin": "TaiAn",
                "name_abbreviation": "ta",
                "code": "3709"
            },
            u"临沂": {
                "name_pinyin": "LinYi",
                "name_abbreviation": "liny",
                "code": "3713"
            },
            u"枣庄": {
                "name_pinyin": "ZaoZhuang",
                "name_abbreviation": "zaoz",
                "code": "3704"
            },
            u"济南": {
                "name_pinyin": "JiNan",
                "name_abbreviation": "jn",
                "code": "3701"
            }
        }
    },
    u"吉林": {
        "name_pinyin": "JiLin",
        "name_abbreviation": "jl",
        "code": "22",
        "city": {
            u"四平": {
                "name_pinyin": "SiPing",
                "name_abbreviation": "sp",
                "code": "2203"
            },
            u"白山": {
                "name_pinyin": "BaiShan",
                "name_abbreviation": "bsh",
                "code": "2206"
            },
            u"辽源": {
                "name_pinyin": "LiaoYuan",
                "name_abbreviation": "liaoy",
                "code": "2204"
            },
            u"松原": {
                "name_pinyin": "SongYuan",
                "name_abbreviation": "songy",
                "code": "2207"
            },
            u"延边": {
                "name_pinyin": "YanBian",
                "name_abbreviation": "yanb",
                "code": "2224"
            },
            u"长春": {
                "name_pinyin": "ChangChun",
                "name_abbreviation": "cc",
                "code": "2201"
            },
            u"白城": {
                "name_pinyin": "BaiCheng",
                "name_abbreviation": "bch",
                "code": "2208"
            },
            u"通化": {
                "name_pinyin": "TongHua",
                "name_abbreviation": "th",
                "code": "2205"
            },
            u"吉林": {
                "name_pinyin": "JiLin",
                "name_abbreviation": "jl",
                "code": "2202"
            }
        }
    },
    u"宁夏": {
        "name_pinyin": "NingXia",
        "name_abbreviation": "nx",
        "code": "64",
        "city": {
            u"固原": {
                "name_pinyin": "GuYuan",
                "name_abbreviation": "guy",
                "code": "6404"
            },
            u"中卫": {
                "name_pinyin": "ZhongWei",
                "name_abbreviation": "zw",
                "code": "6405"
            },
            u"银川": {
                "name_pinyin": "YinChuan",
                "name_abbreviation": "yinc",
                "code": "6401"
            },
            u"吴忠": {
                "name_pinyin": "WuZhong",
                "name_abbreviation": "wzh",
                "code": "6403"
            },
            u"石嘴山": {
                "name_pinyin": "ShiZuiShan",
                "name_abbreviation": "shzsh",
                "code": "6402"
            }
        }
    },
    u"江苏": {
        "name_pinyin": "JiangSu",
        "name_abbreviation": "js",
        "code": "32",
        "city": {
            u"淮安": {
                "name_pinyin": "HuaiAn",
                "name_abbreviation": "ha",
                "code": "3208"
            },
            u"无锡": {
                "name_pinyin": "WuXi",
                "name_abbreviation": "wx",
                "code": "3202"
            },
            u"镇江": {
                "name_pinyin": "ZhenJiang",
                "name_abbreviation": "zhenj",
                "code": "3211"
            },
            u"泰州": {
                "name_pinyin": "TaiZhou",
                "name_abbreviation": "taiz",
                "code": "3212"
            },
            u"南京": {
                "name_pinyin": "NanJing",
                "name_abbreviation": "nj",
                "code": "3201"
            },
            u"徐州": {
                "name_pinyin": "XuZhou",
                "name_abbreviation": "xz",
                "code": "3203"
            },
            u"连云港": {
                "name_pinyin": "LianYunGang",
                "name_abbreviation": "lyg",
                "code": "3207"
            },
            u"苏州": {
                "name_pinyin": "SuZhou",
                "name_abbreviation": "sz",
                "code": "3205"
            },
            u"常州": {
                "name_pinyin": "ChangZhou",
                "name_abbreviation": "changz",
                "code": "3204"
            },
            u"扬州": {
                "name_pinyin": "YangZhou",
                "name_abbreviation": "yz",
                "code": "3210"
            },
            u"盐城": {
                "name_pinyin": "YanCheng",
                "name_abbreviation": "yanc",
                "code": "3209"
            },
            u"南通": {
                "name_pinyin": "NanTong",
                "name_abbreviation": "nt",
                "code": "3206"
            },
            u"宿迁": {
                "name_pinyin": "SuQian",
                "name_abbreviation": "sq",
                "code": "3213"
            },
            u"常熟": {
                "name_pinyin": "ChangShu",
                "name_abbreviation": "changs",
                "code": "3214"
            }
        }
    },
    u"重庆": {
        "city": {
            u"重庆": {
                "name_pinyin": "ChongQing",
                "name_abbreviation": "cq",
                "code": "500000"
            }
        },
        "name_pinyin": "ChongQing",
        "name_abbreviation": "cq",
        "code": "500000"
    },
    u"新疆": {
        "name_pinyin": "XinJiang",
        "name_abbreviation": "xj",
        "code": "65",
        "city": {
            u"乌鲁木齐": {
                "name_pinyin": "WuLuMuQi",
                "name_abbreviation": "wlmq",
                "code": "6501"
            },
            u"喀什": {
                "name_pinyin": "KaShi",
                "name_abbreviation": "ksh",
                "code": "6531"
            },
            u"博尔塔拉": {
                "name_pinyin": "BoErTaLa",
                "name_abbreviation": "betl",
                "code": "6527"
            },
            u"塔城": {
                "name_pinyin": "TaCheng",
                "name_abbreviation": "tch",
                "code": "6542"
            },
            u"伊犁": {
                "name_pinyin": "YiLi",
                "name_abbreviation": "yl",
                "code": "6540"
            },
            u"石河子": {
                "name_pinyin": "ShiHeZi",
                "name_abbreviation": "shhz",
                "code": "6590"
            },
            u"阿克苏": {
                "name_pinyin": "AKeSu",
                "name_abbreviation": "aks",
                "code": "6529"
            },
            u"昌吉": {
                "name_pinyin": "ChangJi",
                "name_abbreviation": "chj",
                "code": "6523"
            },
            u"克拉玛依": {
                "name_pinyin": "KeLaMaYi",
                "name_abbreviation": "klmy",
                "code": "6502"
            },
            u"克孜勒苏柯尔克孜": {
                "name_pinyin": "KeZiLeiSuKeErKeZi",
                "name_abbreviation": "kzlskekz",
                "code": "6530"
            },
            u"和田": {
                "name_pinyin": "HeTian",
                "name_abbreviation": "ht",
                "code": "6532"
            },
            u"吐鲁番": {
                "name_pinyin": "TuLuFan",
                "name_abbreviation": "tlf",
                "code": "6504"
            },
            u"哈密": {
                "name_pinyin": "HaMi",
                "name_abbreviation": "hm",
                "code": "6505"
            },
            u"巴音郭楞": {
                "name_pinyin": "BaYinGuoLeng",
                "name_abbreviation": "bygl",
                "code": "6528"
            },
            u"阿勒泰": {
                "name_pinyin": "ALeiTai",
                "name_abbreviation": "alt",
                "code": "6543"
            }
        }
    },
    u"福建": {
        "name_pinyin": "FuJian",
        "name_abbreviation": "fj",
        "code": "35",
        "city": {
            u"莆田": {
                "name_pinyin": "PuTian",
                "name_abbreviation": "pt",
                "code": "3503"
            },
            u"福州": {
                "name_pinyin": "FuZhou",
                "name_abbreviation": "fzh",
                "code": "3501"
            },
            u"南平": {
                "name_pinyin": "NanPing",
                "name_abbreviation": "np",
                "code": "3507"
            },
            u"漳州": {
                "name_pinyin": "ZhangZhou",
                "name_abbreviation": "zhangz",
                "code": "3506"
            },
            u"宁德": {
                "name_pinyin": "NingDe",
                "name_abbreviation": "nd",
                "code": "3509"
            },
            u"三明": {
                "name_pinyin": "SanMing",
                "name_abbreviation": "sm",
                "code": "3504"
            },
            u"厦门": {
                "name_pinyin": "XiaMen",
                "name_abbreviation": "xm",
                "code": "3502"
            },
            u"龙岩": {
                "name_pinyin": "LongYan",
                "name_abbreviation": "longy",
                "code": "3508"
            },
            u"泉州": {
                "name_pinyin": "QuanZhou",
                "name_abbreviation": "quanz",
                "code": "3505"
            }
        }
    },
    u"浙江": {
        "name_pinyin": "ZheJiang",
        "name_abbreviation": "zhj",
        "code": "33",
        "city": {
            u"丽水": {
                "name_pinyin": "LiShui",
                "name_abbreviation": "lis",
                "code": "3311"
            },
            u"宁波": {
                "name_pinyin": "NingBo",
                "name_abbreviation": "nb",
                "code": "3302"
            },
            u"温州": {
                "name_pinyin": "WenZhou",
                "name_abbreviation": "wz",
                "code": "3303"
            },
            u"台州": {
                "name_pinyin": "TaiZhou",
                "name_abbreviation": "tz",
                "code": "3310"
            },
            u"嘉兴": {
                "name_pinyin": "JiaXing",
                "name_abbreviation": "jx",
                "code": "3304"
            },
            u"舟山": {
                "name_pinyin": "ZhouShan",
                "name_abbreviation": "zhous",
                "code": "3309"
            },
            u"绍兴": {
                "name_pinyin": "ShaoXing",
                "name_abbreviation": "sx",
                "code": "3306"
            },
            u"衢州": {
                "name_pinyin": "QuZhou",
                "name_abbreviation": "qz",
                "code": "3308"
            },
            u"杭州": {
                "name_pinyin": "HangZhou",
                "name_abbreviation": "hz",
                "code": "3301"
            },
            u"湖州": {
                "name_pinyin": "HuZhou",
                "name_abbreviation": "huz",
                "code": "3305"
            },
            u"金华": {
                "name_pinyin": "JinHua",
                "name_abbreviation": "jh",
                "code": "3307"
            }
        }
    },
    u"湖北": {
        "name_pinyin": "HuBei",
        "name_abbreviation": "hb",
        "code": "42",
        "city": {
            u"仙桃": {
                "name_pinyin": "XianTao",
                "name_abbreviation": "xiant",
                "code": "4290"
            },
            u"恩施": {
                "name_pinyin": "EnShi",
                "name_abbreviation": "esh",
                "code": "4228"
            },
            u"黄冈": {
                "name_pinyin": "HuangGang",
                "name_abbreviation": "huangg",
                "code": "4211"
            },
            u"随州": {
                "name_pinyin": "SuiZhou",
                "name_abbreviation": "suizh",
                "code": "4213"
            },
            u"孝感": {
                "name_pinyin": "XiaoGan",
                "name_abbreviation": "xiaog",
                "code": "4209"
            },
            u"咸宁": {
                "name_pinyin": "XianNing",
                "name_abbreviation": "xiann",
                "code": "4212"
            },
            u"荆州": {
                "name_pinyin": "JingZhou",
                "name_abbreviation": "jingz",
                "code": "4210"
            },
            u"襄阳": {
                "name_pinyin": "XiangYang",
                "name_abbreviation": "xy",
                "code": "4206"
            },
            u"十堰": {
                "name_pinyin": "ShiYan",
                "name_abbreviation": "shiy",
                "code": "4203"
            },
            u"荆门": {
                "name_pinyin": "JingMen",
                "name_abbreviation": "jiangm",
                "code": "4208"
            },
            u"武汉": {
                "name_pinyin": "WuHan",
                "name_abbreviation": "wh",
                "code": "4201"
            },
            u"宜昌": {
                "name_pinyin": "YiChang",
                "name_abbreviation": "yic",
                "code": "4205"
            },
            u"黄石": {
                "name_pinyin": "HuangShi",
                "name_abbreviation": "huangs",
                "code": "4202"
            },
            u"鄂州": {
                "name_pinyin": "EZhou",
                "name_abbreviation": "ezh",
                "code": "4207"
            }
        }
    },
    u"天津": {
        "name_pinyin": "TianJin",
        "city": {
            u"天津": {
                "name_pinyin": "TianJin",
                "name_abbreviation": "tj",
                "code": "120000"
            }
        },
        "name_abbreviation": "tj",
        "code": "120000"
    },
    u"江西": {
        "name_pinyin": "JiangXi",
        "name_abbreviation": "jx",
        "code": "36",
        "city": {
            u"新余": {
                "name_pinyin": "XinYu",
                "name_abbreviation": "xinyu",
                "code": "3605"
            },
            u"萍乡": {
                "name_pinyin": "PingXiang",
                "name_abbreviation": "px",
                "code": "3603"
            },
            u"景德镇": {
                "name_pinyin": "JingDeZhen",
                "name_abbreviation": "jdz",
                "code": "3602"
            },
            u"南昌": {
                "name_pinyin": "NanChang",
                "name_abbreviation": "nc",
                "code": "3601"
            },
            u"吉安": {
                "name_pinyin": "JiAn",
                "name_abbreviation": "ja",
                "code": "3608"
            },
            u"抚州": {
                "name_pinyin": "FuZhou",
                "name_abbreviation": "fuz",
                "code": "3610"
            },
            u"宜春": {
                "name_pinyin": "YiChun",
                "name_abbreviation": "yich",
                "code": "3609"
            },
            u"赣州": {
                "name_pinyin": "GanZhou",
                "name_abbreviation": "ganz",
                "code": "3607"
            },
            u"上饶": {
                "name_pinyin": "ShangRao",
                "name_abbreviation": "sr",
                "code": "3611"
            },
            u"九江": {
                "name_pinyin": "JiuJiang",
                "name_abbreviation": "jj",
                "code": "3604"
            },
            u"鹰潭": {
                "name_pinyin": "YingTan",
                "name_abbreviation": "yingt",
                "code": "3606"
            }
        }
    },
    u"西藏": {
        "name_pinyin": "XiZang",
        "name_abbreviation": "xz",
        "code": "54",
        "city": {
            u"拉萨": {
                "name_pinyin": "LaSa",
                "name_abbreviation": "las",
                "code": "5401"
            },
            u"阿里": {
                "name_pinyin": "ALi",
                "name_abbreviation": "al",
                "code": "5425"
            },
            u"山南": {
                "name_pinyin": "ShanNan",
                "name_abbreviation": "shann",
                "code": "5405"
            },
            u"日喀则": {
                "name_pinyin": "RiKaZe",
                "name_abbreviation": "rkz",
                "code": "5402"
            },
            u"那曲": {
                "name_pinyin": "NaQu",
                "name_abbreviation": "nq",
                "code": "5406"
            },
            u"林芝": {
                "name_pinyin": "LinZhi",
                "name_abbreviation": "lzh",
                "code": "5404"
            },
            u"昌都": {
                "name_pinyin": "ChangDou",
                "name_abbreviation": "chd",
                "code": "5403"
            }
        }
    },
    u"黑龙江": {
        "name_pinyin": "HeiLongJiang",
        "name_abbreviation": "hlj",
        "code": "23",
        "city": {
            u"黑河": {
                "name_pinyin": "HeiHe",
                "name_abbreviation": "hh",
                "code": "2311"
            },
            u"绥化": {
                "name_pinyin": "SuiHua",
                "name_abbreviation": "suih",
                "code": "2312"
            },
            u"齐齐哈尔": {
                "name_pinyin": "QiQiHaEr",
                "name_abbreviation": "qqhe",
                "code": "2302"
            },
            u"双鸭山": {
                "name_pinyin": "ShuangYaShan",
                "name_abbreviation": "shysh",
                "code": "2305"
            },
            u"大兴安岭": {
                "name_pinyin": "DaXingAnLing",
                "name_abbreviation": "dxal",
                "code": "2327"
            },
            u"哈尔滨": {
                "name_pinyin": "HaErBin",
                "name_abbreviation": "hrb",
                "code": "2301"
            },
            u"大庆": {
                "name_pinyin": "DaQing",
                "name_abbreviation": "dq",
                "code": "2306"
            },
            u"牡丹江": {
                "name_pinyin": "MuDanJiang",
                "name_abbreviation": "mdj",
                "code": "2310"
            },
            u"伊春": {
                "name_pinyin": "YiChun",
                "name_abbreviation": "yc",
                "code": "2307"
            },
            u"鹤岗": {
                "name_pinyin": "HeGang",
                "name_abbreviation": "hg",
                "code": "2304"
            },
            u"七台河": {
                "name_pinyin": "QiTaiHe",
                "name_abbreviation": "qth",
                "code": "2309"
            },
            u"鸡西": {
                "name_pinyin": "JiXi",
                "name_abbreviation": "jix",
                "code": "2303"
            },
            u"佳木斯": {
                "name_pinyin": "JiaMuSi",
                "name_abbreviation": "jms",
                "code": "2308"
            }
        }
    },
    u"广东": {
        "name_pinyin": "GuangDong",
        "name_abbreviation": "gd",
        "code": "44",
        "city": {
            u"广州": {
                "name_pinyin": "GuangZhou",
                "name_abbreviation": "gz",
                "code": "4401"
            },
            u"深圳": {
                "name_pinyin": "ShenZhen",
                "name_abbreviation": "shenz",
                "code": "4403"
            },
            u"汕尾": {
                "name_pinyin": "ShanWei",
                "name_abbreviation": "shanw",
                "code": "4415"
            },
            u"阳江": {
                "name_pinyin": "YangJiang",
                "name_abbreviation": "yj",
                "code": "4417"
            },
            u"潮州": {
                "name_pinyin": "ChaoZhou",
                "name_abbreviation": "chzh",
                "code": "4451"
            },
            u"清远": {
                "name_pinyin": "QingYuan",
                "name_abbreviation": "qingy",
                "code": "4418"
            },
            u"东莞": {
                "name_pinyin": "DongGuan",
                "name_abbreviation": "dg",
                "code": "4419"
            },
            u"茂名": {
                "name_pinyin": "MaoMing",
                "name_abbreviation": "mm",
                "code": "4409"
            },
            u"河源": {
                "name_pinyin": "HeYuan",
                "name_abbreviation": "hey",
                "code": "4416"
            },
            u"湛江": {
                "name_pinyin": "ZhanJiang",
                "name_abbreviation": "zhanj",
                "code": "4408"
            },
            u"佛山": {
                "name_pinyin": "FoShan",
                "name_abbreviation": "fs",
                "code": "4406"
            },
            u"汕头": {
                "name_pinyin": "ShanTou",
                "name_abbreviation": "st",
                "code": "4405"
            },
            u"肇庆": {
                "name_pinyin": "ZhaoQing",
                "name_abbreviation": "zq",
                "code": "4412"
            },
            u"韶关": {
                "name_pinyin": "ShaoGuan",
                "name_abbreviation": "shg",
                "code": "4402"
            },
            u"中山": {
                "name_pinyin": "ZhongShan",
                "name_abbreviation": "zs",
                "code": "4420"
            },
            u"珠海": {
                "name_pinyin": "ZhuHai",
                "name_abbreviation": "zhh",
                "code": "4404"
            },
            u"梅州": {
                "name_pinyin": "MeiZhou",
                "name_abbreviation": "mzh",
                "code": "4414"
            },
            u"江门": {
                "name_pinyin": "JiangMen",
                "name_abbreviation": "jm",
                "code": "4407"
            },
            u"云浮": {
                "name_pinyin": "YunFu",
                "name_abbreviation": "yf",
                "code": "4453"
            },
            u"惠州": {
                "name_pinyin": "HuiZhou",
                "name_abbreviation": "huiz",
                "code": "4413"
            },
            u"揭阳": {
                "name_pinyin": "JieYang",
                "name_abbreviation": "jiey",
                "code": "4452"
            }
        }
    },
    u"安徽": {
        "name_pinyin": "AnHui",
        "name_abbreviation": "ah",
        "code": "34",
        "city": {
            u"芜湖": {
                "name_pinyin": "WuHu",
                "name_abbreviation": "wuh",
                "code": "3402"
            },
            u"宿州": {
                "name_pinyin": "SuZhou",
                "name_abbreviation": "suz",
                "code": "3413"
            },
            u"铜陵": {
                "name_pinyin": "TongLing",
                "name_abbreviation": "tongling",
                "code": "3407"
            },
            u"阜阳": {
                "name_pinyin": "FuYang",
                "name_abbreviation": "fy",
                "code": "3412"
            },
            u"蚌埠": {
                "name_pinyin": "BengBu",
                "name_abbreviation": "bengb",
                "code": "3403"
            },
            u"马鞍山": {
                "name_pinyin": "MaAnShan",
                "name_abbreviation": "mas",
                "code": "3405"
            },
            u"六安": {
                "name_pinyin": "LiuAn",
                "name_abbreviation": "liua",
                "code": "3415"
            },
            u"宣城": {
                "name_pinyin": "XuanCheng",
                "name_abbreviation": "xch",
                "code": "3418"
            },
            u"安庆": {
                "name_pinyin": "AnQing",
                "name_abbreviation": "aq",
                "code": "3408"
            },
            u"淮南": {
                "name_pinyin": "HuaiNan",
                "name_abbreviation": "hn",
                "code": "3404"
            },
            u"亳州": {
                "name_pinyin": "BoZhou",
                "name_abbreviation": "bozh",
                "code": "3416"
            },
            u"滁州": {
                "name_pinyin": "ChuZhou",
                "name_abbreviation": "chuz",
                "code": "3411"
            },
            u"淮北": {
                "name_pinyin": "HuaiBei",
                "name_abbreviation": "huaib",
                "code": "3406"
            },
            u"池州": {
                "name_pinyin": "ChiZhou",
                "name_abbreviation": "chiz",
                "code": "3417"
            },
            u"合肥": {
                "name_pinyin": "HeFei",
                "name_abbreviation": "hf",
                "code": "3401"
            },
            u"黄山": {
                "name_pinyin": "HuangShan",
                "name_abbreviation": "hsh",
                "code": "3410"
            }
        }
    },
    u"内蒙古": {
        "name_pinyin": "NeiMengGu",
        "name_abbreviation": "nmg",
        "code": "15",
        "city": {
            u"兴安盟": {
                "name_pinyin": "XingAnMeng",
                "name_abbreviation": "xam",
                "code": "1522"
            },
            u"乌海": {
                "name_pinyin": "WuHai",
                "name_abbreviation": "whai",
                "code": "1503"
            },
            u"呼伦贝尔": {
                "name_pinyin": "HuLunBeiEr",
                "name_abbreviation": "hlbe",
                "code": "1507"
            },
            u"乌兰察布": {
                "name_pinyin": "WuLanChaBu",
                "name_abbreviation": "wlcb",
                "code": "1509"
            },
            u"鄂尔多斯": {
                "name_pinyin": "EErDuoSi",
                "name_abbreviation": "erds",
                "code": "1506"
            },
            u"通辽": {
                "name_pinyin": "TongLiao",
                "name_abbreviation": "tongl",
                "code": "1505"
            },
            u"阿拉善盟": {
                "name_pinyin": "ALaShanMeng",
                "name_abbreviation": "alshm",
                "code": "1529"
            },
            u"赤峰": {
                "name_pinyin": "ChiFeng",
                "name_abbreviation": "chf",
                "code": "1504"
            },
            u"锡林郭勒盟": {
                "name_pinyin": "XiLinGuoLeiMeng",
                "name_abbreviation": "xlglm",
                "code": "1525"
            },
            u"巴彦淖尔": {
                "name_pinyin": "BaYanNaoEr",
                "name_abbreviation": "byne",
                "code": "1508"
            },
            u"包头": {
                "name_pinyin": "BaoTou",
                "name_abbreviation": "bt",
                "code": "1502"
            },
            u"呼和浩特": {
                "name_pinyin": "HuHeHaoTe",
                "name_abbreviation": "hhht",
                "code": "1501"
            }
        }
    },
    u"云南": {
        "name_pinyin": "YunNan",
        "name_abbreviation": "yn",
        "code": "53",
        "city": {
            u"曲靖": {
                "name_pinyin": "QuJing",
                "name_abbreviation": "qj",
                "code": "5303"
            },
            u"怒江": {
                "name_pinyin": "NuJiang",
                "name_abbreviation": "nuj",
                "code": "5333"
            },
            u"保山": {
                "name_pinyin": "BaoShan",
                "name_abbreviation": "bsh",
                "code": "5305"
            },
            u"西双版纳": {
                "name_pinyin": "XiShuangBanNa",
                "name_abbreviation": "xshbn",
                "code": "5328"
            },
            u"文山": {
                "name_pinyin": "WenShan",
                "name_abbreviation": "wsh",
                "code": "5326"
            },
            u"楚雄": {
                "name_pinyin": "ChuXiong",
                "name_abbreviation": "chx",
                "code": "5323"
            },
            u"丽江": {
                "name_pinyin": "LiJiang",
                "name_abbreviation": "lj",
                "code": "5307"
            },
            u"红河": {
                "name_pinyin": "HongHe",
                "name_abbreviation": "hongh",
                "code": "5325"
            },
            u"大理": {
                "name_pinyin": "DaLi",
                "name_abbreviation": "dal",
                "code": "5329"
            },
            u"玉溪": {
                "name_pinyin": "YuXi",
                "name_abbreviation": "yux",
                "code": "5304"
            },
            u"普洱": {
                "name_pinyin": "PuEr",
                "name_abbreviation": "pe",
                "code": "5308"
            },
            u"临沧": {
                "name_pinyin": "LinCang",
                "name_abbreviation": "lc",
                "code": "5309"
            },
            u"迪庆": {
                "name_pinyin": "DiQing",
                "name_abbreviation": "diq",
                "code": "5334"
            },
            u"德宏": {
                "name_pinyin": "DeHong",
                "name_abbreviation": "dh",
                "code": "5331"
            },
            u"昆明": {
                "name_pinyin": "KunMing",
                "name_abbreviation": "km",
                "code": "5301"
            },
            u"昭通": {
                "name_pinyin": "ZhaoTong",
                "name_abbreviation": "zt",
                "code": "5306"
            }
        }
    },
    u"香港": {
        "name_pinyin": "XiangGang",
        "name_abbreviation": "xg",
        "code": "81",
        "city": {
            u"香港": {
                "name_pinyin": "XiangGang",
                "name_abbreviation": "xg",
                "code": "81"
            }
        }
    },
    u"北京": {
        "name_pinyin": "BeiJing",
        "city": {
            u"北京": {
                "name_pinyin": "BeiJing",
                "name_abbreviation": "bj",
                "code": "110000"
            }
        },
        "name_abbreviation": "bj",
        "code": "110000"
    },
    u"台湾": {
        "name_pinyin": "TaiWan",
        "name_abbreviation": "tw",
        "code": "71",
        "city": {}
    },
    u"贵州": {
        "name_pinyin": "GuiZhou",
        "name_abbreviation": "gzh",
        "code": "52",
        "city": {
            u"黔西南": {
                "name_pinyin": "QianXiNan",
                "name_abbreviation": "qxn",
                "code": "5223"
            },
            u"黔东南": {
                "name_pinyin": "QianDongNan",
                "name_abbreviation": "qdn",
                "code": "5226"
            },
            u"安顺": {
                "name_pinyin": "AnShun",
                "name_abbreviation": "ash",
                "code": "5204"
            },
            u"遵义": {
                "name_pinyin": "ZunYi",
                "name_abbreviation": "zy",
                "code": "5203"
            },
            u"铜仁": {
                "name_pinyin": "TongRen",
                "name_abbreviation": "tr",
                "code": "5206"
            },
            u"黔南": {
                "name_pinyin": "QianNan",
                "name_abbreviation": "qn",
                "code": "5227"
            },
            u"六盘水": {
                "name_pinyin": "LiuPanShui",
                "name_abbreviation": "lps",
                "code": "5202"
            },
            u"贵阳": {
                "name_pinyin": "GuiYang",
                "name_abbreviation": "gy",
                "code": "5201"
            },
            u"毕节": {
                "name_pinyin": "BiJie",
                "name_abbreviation": "bij",
                "code": "5205"
            },
            u"贵安": {
                "name_pinyin": "GuiAn",
                "name_abbreviation": "ga",
                "code": "5228"
            },
            u"兴义": {
                "name_pinyin": "XingYi",
                "name_abbreviation": "xingy",
                "code": "5229"
            }
        }
    },
    u"山西": {
        "name_pinyin": "ShanXi",
        "name_abbreviation": "shx",
        "code": "14",
        "city": {
            u"朔州": {
                "name_pinyin": "ShuoZhou",
                "name_abbreviation": "shuoz",
                "code": "1406"
            },
            u"临汾": {
                "name_pinyin": "LinFen",
                "name_abbreviation": "lf",
                "code": "1410"
            },
            u"太原": {
                "name_pinyin": "TaiYuan",
                "name_abbreviation": "ty",
                "code": "1401"
            },
            u"长治": {
                "name_pinyin": "ChangZhi",
                "name_abbreviation": "changzh",
                "code": "1404"
            },
            u"吕梁": {
                "name_pinyin": "LvLiang",
                "name_abbreviation": "lvl",
                "code": "1411"
            },
            u"晋城": {
                "name_pinyin": "JinCheng",
                "name_abbreviation": "jinch",
                "code": "1405"
            },
            u"晋中": {
                "name_pinyin": "JinZhong",
                "name_abbreviation": "jz",
                "code": "1407"
            },
            u"忻州": {
                "name_pinyin": "XinZhou",
                "name_abbreviation": "xzh",
                "code": "1409"
            },
            u"阳泉": {
                "name_pinyin": "YangQuan",
                "name_abbreviation": "yq",
                "code": "1403"
            },
            u"大同": {
                "name_pinyin": "DaTong",
                "name_abbreviation": "dt",
                "code": "1402"
            },
            u"运城": {
                "name_pinyin": "YunCheng",
                "name_abbreviation": "yunc",
                "code": "1408"
            }
        }
    },
    u"海南": {
        "name_pinyin": "HaiNan",
        "name_abbreviation": "hn",
        "code": "46",
        "city": {
            u"三沙": {
                "name_pinyin": "SanSha",
                "name_abbreviation": "sans",
                "code": "4603"
            },
            u"三亚": {
                "name_pinyin": "SanYa",
                "name_abbreviation": "sany",
                "code": "4602"
            },
            u"海口": {
                "name_pinyin": "HaiKou",
                "name_abbreviation": "hk",
                "code": "4601"
            },
            u"儋州": {
                "name_pinyin": "DanZhou",
                "name_abbreviation": "dzh",
                "code": "4604"
            },
            u"五指山": {
                "name_pinyin": "WuZhiShan",
                "name_abbreviation": "wzs",
                "code": "4690"
            }
        }
    },
    u"澳门": {
        "name_pinyin": "AoMen",
        "name_abbreviation": "am",
        "code": "82",
        "city": {
            u"澳门": {
                "name_pinyin": "AoMen",
                "name_abbreviation": "am",
                "code": "82"
            }
        }
    },
    u"辽宁": {
        "name_pinyin": "LiaoNing",
        "name_abbreviation": "ln",
        "code": "21",
        "city": {
            u"铁岭": {
                "name_pinyin": "TieLing",
                "name_abbreviation": "tl",
                "code": "2112"
            },
            u"辽阳": {
                "name_pinyin": "LiaoYang",
                "name_abbreviation": "liaoy",
                "code": "2110"
            },
            u"阜新": {
                "name_pinyin": "FuXin",
                "name_abbreviation": "fx",
                "code": "2109"
            },
            u"大连": {
                "name_pinyin": "DaLian",
                "name_abbreviation": "dl",
                "code": "2102"
            },
            u"锦州": {
                "name_pinyin": "JinZhou",
                "name_abbreviation": "jinz",
                "code": "2107"
            },
            u"本溪": {
                "name_pinyin": "BenXi",
                "name_abbreviation": "bx",
                "code": "2105"
            },
            u"葫芦岛": {
                "name_pinyin": "HuLuDao",
                "name_abbreviation": "hld",
                "code": "2114"
            },
            u"营口": {
                "name_pinyin": "YingKou",
                "name_abbreviation": "yink",
                "code": "2108"
            },
            u"盘锦": {
                "name_pinyin": "PanJin",
                "name_abbreviation": "pj",
                "code": "2111"
            },
            u"鞍山": {
                "name_pinyin": "AnShan",
                "name_abbreviation": "as",
                "code": "2103"
            },
            u"丹东": {
                "name_pinyin": "DanDong",
                "name_abbreviation": "dd",
                "code": "2106"
            },
            u"抚顺": {
                "name_pinyin": "FuShun",
                "name_abbreviation": "fus",
                "code": "2104"
            },
            u"沈阳": {
                "name_pinyin": "ShenYang",
                "name_abbreviation": "sy",
                "code": "2101"
            },
            u"朝阳": {
                "name_pinyin": "ChaoYang",
                "name_abbreviation": "chaoy",
                "code": "2113"
            }
        }
    },
    u"青海": {
        "name_pinyin": "QingHai",
        "name_abbreviation": "qh",
        "code": "63",
        "city": {
            u"海北": {
                "name_pinyin": "HaiBei",
                "name_abbreviation": "haib",
                "code": "6322"
            },
            u"海南": {
                "name_pinyin": "HaiNan",
                "name_abbreviation": "hain",
                "code": "6325"
            },
            u"玉树": {
                "name_pinyin": "YuShu",
                "name_abbreviation": "ysh",
                "code": "6327"
            },
            u"黄南": {
                "name_pinyin": "HuangNan",
                "name_abbreviation": "huangn",
                "code": "6323"
            },
            u"海东": {
                "name_pinyin": "HaiDong",
                "name_abbreviation": "haid",
                "code": "6302"
            },
            u"海西": {
                "name_pinyin": "HaiXi",
                "name_abbreviation": "hx",
                "code": "6328"
            },
            u"果洛": {
                "name_pinyin": "GuoLuo",
                "name_abbreviation": "guol",
                "code": "6326"
            },
            u"西宁": {
                "name_pinyin": "XiNing",
                "name_abbreviation": "xin",
                "code": "6301"
            }
        }
    },
    u"甘肃": {
        "name_pinyin": "GanSu",
        "name_abbreviation": "gs",
        "code": "62",
        "city": {
            u"陇南": {
                "name_pinyin": "LongNan",
                "name_abbreviation": "ln",
                "code": "6212"
            },
            u"嘉峪关": {
                "name_pinyin": "JiaYuGuan",
                "name_abbreviation": "jyg",
                "code": "6202"
            },
            u"临夏": {
                "name_pinyin": "LinXia",
                "name_abbreviation": "lx",
                "code": "6229"
            },
            u"定西": {
                "name_pinyin": "DingXi",
                "name_abbreviation": "dx",
                "code": "6211"
            },
            u"兰州": {
                "name_pinyin": "LanZhou",
                "name_abbreviation": "lz",
                "code": "6201"
            },
            u"天水": {
                "name_pinyin": "TianShui",
                "name_abbreviation": "tians",
                "code": "6205"
            },
            u"金昌": {
                "name_pinyin": "JinChang",
                "name_abbreviation": "jch",
                "code": "6203"
            },
            u"武威": {
                "name_pinyin": "WuWei",
                "name_abbreviation": "ww",
                "code": "6206"
            },
            u"平凉": {
                "name_pinyin": "PingLiang",
                "name_abbreviation": "pl",
                "code": "6208"
            },
            u"庆阳": {
                "name_pinyin": "QingYang",
                "name_abbreviation": "qy",
                "code": "6210"
            },
            u"张掖": {
                "name_pinyin": "ZhangYe",
                "name_abbreviation": "zhangy",
                "code": "6207"
            },
            u"甘南": {
                "name_pinyin": "GanNan",
                "name_abbreviation": "gn",
                "code": "6230"
            },
            u"酒泉": {
                "name_pinyin": "JiuQuan",
                "name_abbreviation": "jq",
                "code": "6209"
            },
            u"白银": {
                "name_pinyin": "BaiYin",
                "name_abbreviation": "by",
                "code": "6204"
            }
        }
    },
    u"四川": {
        "name_pinyin": "SiChuan",
        "name_abbreviation": "sch",
        "code": "51",
        "city": {
            u"凉山": {
                "name_pinyin": "LiangShan",
                "name_abbreviation": "liangs",
                "code": "5134"
            },
            u"南充": {
                "name_pinyin": "NanChong",
                "name_abbreviation": "nanc",
                "code": "5113"
            },
            u"乐山": {
                "name_pinyin": "LeShan",
                "name_abbreviation": "les",
                "code": "5111"
            },
            u"雅安": {
                "name_pinyin": "YaAn",
                "name_abbreviation": "ya",
                "code": "5118"
            },
            u"攀枝花": {
                "name_pinyin": "PanZhiHua",
                "name_abbreviation": "pzh",
                "code": "5104"
            },
            u"达州": {
                "name_pinyin": "DaZhou",
                "name_abbreviation": "dz",
                "code": "5117"
            },
            u"眉山": {
                "name_pinyin": "MeiShan",
                "name_abbreviation": "msh",
                "code": "5114"
            },
            u"资阳": {
                "name_pinyin": "ZiYang",
                "name_abbreviation": "ziy",
                "code": "5120"
            },
            u"阿坝": {
                "name_pinyin": "ABa",
                "name_abbreviation": "ab",
                "code": "5132"
            },
            u"绵阳": {
                "name_pinyin": "MianYang",
                "name_abbreviation": "my",
                "code": "5107"
            },
            u"成都": {
                "name_pinyin": "ChengDu",
                "name_abbreviation": "cd",
                "code": "5101"
            },
            u"宜宾": {
                "name_pinyin": "YiBin",
                "name_abbreviation": "yb",
                "code": "5115"
            },
            u"内江": {
                "name_pinyin": "NeiJiang",
                "name_abbreviation": "neij",
                "code": "5110"
            },
            u"巴中": {
                "name_pinyin": "BaZhong",
                "name_abbreviation": "bzh",
                "code": "5119"
            },
            u"德阳": {
                "name_pinyin": "DeYang",
                "name_abbreviation": "dy",
                "code": "5106"
            },
            u"甘孜": {
                "name_pinyin": "GanZi",
                "name_abbreviation": "ganzi",
                "code": "5133"
            },
            u"广安": {
                "name_pinyin": "GuangAn",
                "name_abbreviation": "guanga",
                "code": "5116"
            },
            u"泸州": {
                "name_pinyin": "LuZhou",
                "name_abbreviation": "luz",
                "code": "5105"
            },
            u"广元": {
                "name_pinyin": "GuangYuan",
                "name_abbreviation": "guangy",
                "code": "5108"
            },
            u"自贡": {
                "name_pinyin": "ZiGong",
                "name_abbreviation": "zg",
                "code": "5103"
            },
            u"遂宁": {
                "name_pinyin": "SuiNing",
                "name_abbreviation": "sn",
                "code": "5109"
            }
        }
    }
}
"""
__city__ = {
    u"广西": {
        "name_pinyin": "GuangXi",
        "name_abbreviation": "gx",
        "code": "45",
        "city": {
            u"防城港": {
                "name_pinyin": "FangChengGang",
                "name_abbreviation": "fcg",
                "code": "4506"
            },
        }
    },
"""
if __name__ == '__main__':
    result = {}
    for raw_key, raw_val in raw_city.items():
        raw_key = raw_key.replace('自治区', '').replace('特别行政区', '').replace('维吾尔', '').replace('回族', '').replace('区', '').replace('市', '').replace('省', '').replace('壮族', '')
        if raw_key not in __city__:
            print(raw_key)
        else:
            new_city = raw_val['city']
            old_city = __city__[raw_key]['city']
            if raw_key not in result:
                result[raw_key] = {'city': {},
                                   'name_abbreviation': __city__[raw_key]['name_abbreviation'],
                                   'name_pinyin': __city__[raw_key]['name_pinyin'],
                                   'code': raw_val['code']}
            for new_city_key, new_city_val in new_city.items():
                new_city_key = new_city_key.replace('市', '').replace('自治州', '')
                if new_city_key not in old_city:
                    for oc in old_city.keys():
                        if oc in new_city_key:
                            new_city_key = oc
                            break
                    else:
                        print(f'err: {new_city_key}')
                result[raw_key]['city'][new_city_key] = {'name_abbreviation': __city__[raw_key]['city'][new_city_key]['name_abbreviation'],
                                                         'name_pinyin': __city__[raw_key]['city'][new_city_key]['name_pinyin'],
                                                         'code': new_city_val['code']}
    # pprint.pprint(result)
    # print(json.dumps(result))

    # f = Workbook(encoding='utf-8')
    # table = f.add_sheet('data')

    col = 1
    row = 1
    wb = openpyxl.load_workbook('/Users/colin/Desktop/1.xlsx')
    work_sheet = wb['Sheet1']

    for province, pro_val in result.items():
        if pro_val['city']:
            for city, ci_val in pro_val['city'].items():
                # print(f'{province} {city} {ci_val["name_abbreviation"]} {ci_val["name_pinyin"]} {ci_val["code"]}')
                work_sheet.cell(row=row, column=col).value = province
                work_sheet.cell(row=row, column=col+1).value = city
                work_sheet.cell(row=row, column=col+2).value = ci_val["name_abbreviation"]
                work_sheet.cell(row=row, column=col+3).value = ci_val["name_pinyin"]
                work_sheet.cell(row=row, column=col+4).value = ci_val["code"]
                row += 1

                # print(f'{province}')
        else:
            # print(f'{province} {province} {pro_val["name_abbreviation"]} {pro_val["name_pinyin"]} {pro_val["code"]}')
            work_sheet.cell(row=row, column=col).value = province
            work_sheet.cell(row=row, column=col + 1).value = province
            work_sheet.cell(row=row, column=col + 2).value = pro_val["name_abbreviation"]
            work_sheet.cell(row=row, column=col + 3).value = pro_val["name_pinyin"]
            work_sheet.cell(row=row, column=col + 4).value = pro_val["code"]
            row += 1
            # print(f'{province}')
    wb.save('/Users/colin/Desktop/2.xlsx')
